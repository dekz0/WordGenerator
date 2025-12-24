"""
Загрузчик данных из Excel.
Single Responsibility: только загрузка данных.
Используем openpyxl напрямую для минимального размера exe.
"""

from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook


class DataLoaderError(Exception):
    """Исключение при загрузке данных."""

    pass


class DataLoaderProtocol(Protocol):
    """Интерфейс загрузчика данных (Open/Closed principle)."""

    def load(self, file_path: Path) -> list[dict]: ...
    def get_columns(self, file_path: Path) -> list[str]: ...


class ExcelLoader:
    """
    Загрузчик Excel файлов.
    Преобразует данные в список словарей для дальнейшей обработки.
    Использует openpyxl напрямую (без pandas/numpy) для лёгкого exe.
    """

    SUPPORTED_EXTENSIONS = (".xlsx",)

    def load(self, file_path: Path) -> list[dict]:
        """
        Загрузить данные из Excel файла.

        Args:
            file_path: Путь к Excel файлу

        Returns:
            Список словарей, где ключи - названия колонок

        Raises:
            DataLoaderError: При ошибке загрузки
        """
        self._validate_file(file_path)

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                raise DataLoaderError("Excel файл пуст или не содержит данных")

            # Первая строка — заголовки
            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(rows[0])
            ]

            # Остальные строки — данные
            records = []
            for row in rows[1:]:
                # Пропускаем полностью пустые строки
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # Преобразуем в строку, None → пустая строка
                        record[headers[i]] = (
                            str(value).strip() if value is not None else ""
                        )
                records.append(record)

            if not records:
                raise DataLoaderError(
                    "Excel файл не содержит данных (только заголовки)"
                )

            return records

        except DataLoaderError:
            raise
        except Exception as e:
            raise DataLoaderError(f"Ошибка чтения Excel файла: {e}")

    def get_columns(self, file_path: Path) -> list[str]:
        """
        Получить список колонок из Excel файла.

        Args:
            file_path: Путь к Excel файлу

        Returns:
            Список названий колонок
        """
        self._validate_file(file_path)

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            # Читаем только первую строку
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()

            return [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(first_row)
            ]
        except Exception as e:
            raise DataLoaderError(f"Ошибка чтения заголовков Excel: {e}")

    def _validate_file(self, file_path: Path) -> None:
        """Проверить корректность файла."""
        if not file_path.exists():
            raise DataLoaderError(f"Файл не найден: {file_path}")

        if file_path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise DataLoaderError(
                f"Неподдерживаемый формат файла: {file_path.suffix}. "
                f"Поддерживается только .xlsx"
            )

        if file_path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise DataLoaderError(
                f"Неподдерживаемый формат файла: {file_path.suffix}. "
                f"Поддерживаются: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )
